"""Prepare deterministic local assets for the Issue 12 prototype."""

from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _relative(path: Path, root: Path) -> str:
    return path.resolve().relative_to(root.resolve()).as_posix()


def _image_cell(image: Any) -> str:
    marker = getattr(image.anchor, "_from", None)
    if marker is None:
        raise ValueError("unsupported workbook image anchor")
    return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"


def extract_style_images(
    workbook_path: Path,
    sheet_name: str,
    cells: list[str],
    output_dir: Path,
    record_root: Path,
) -> list[dict[str, Any]]:
    """Extract exact embedded image bytes by cell anchor."""
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"missing style sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    images_by_cell: dict[str, list[Any]] = {}
    for image in sheet._images:  # openpyxl exposes drawings only via this API.
        images_by_cell.setdefault(_image_cell(image), []).append(image)

    output_dir.mkdir(parents=True, exist_ok=True)
    records = []
    for cell in cells:
        matches = images_by_cell.get(cell, [])
        if len(matches) != 1:
            raise ValueError(f"style cell {cell} has {len(matches)} embedded images")
        payload = matches[0]._data()
        with Image.open(io.BytesIO(payload)) as decoded:
            decoded.load()
            width, height = decoded.size
            image_format = decoded.format
            mode = decoded.mode
        suffix = f".{(image_format or 'png').lower()}"
        destination = output_dir / f"{cell}{suffix}"
        destination.write_bytes(payload)
        row = matches[0].anchor._from.row + 1
        records.append(
            {
                "cell": cell,
                "row": row,
                "style_name": str(sheet.cell(row, 1).value),
                "game_name": str(sheet.cell(row, 2).value),
                "keywords": str(sheet.cell(row, 3).value),
                "description": str(sheet.cell(row, 4).value),
                "path": _relative(destination, record_root),
                "embedded_sha256": sha256_bytes(payload),
                "output_sha256": sha256_file(destination),
                "format": image_format,
                "mode": mode,
                "width": width,
                "height": height,
            }
        )
    workbook.close()
    return records


def read_compositions(
    workbook_path: Path,
    sheet_name: str,
    rows: list[int],
) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"missing composition sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    records = []
    for row in rows:
        values = [sheet.cell(row, column).value for column in range(1, 8)]
        if any(value is None for value in values):
            raise ValueError(f"composition row {row} is incomplete")
        records.append(
            {
                "row": row,
                "number": str(values[0]),
                "group": str(values[1]),
                "name": str(values[2]),
                "film": str(values[3]),
                "features": str(values[4]),
                "fit": str(values[5]),
                "prompt": str(values[6]),
            }
        )
    workbook.close()
    return records


def split_character_sheet(
    source_path: Path,
    output_dir: Path,
    record_root: Path,
    source_root: Path | None = None,
) -> dict[str, Any]:
    """Split at H//2; an odd extra row belongs to the bottom image."""
    output_dir.mkdir(parents=True, exist_ok=True)
    with Image.open(source_path) as source:
        source.load()
        width, height = source.size
        midpoint = height // 2
        top = source.crop((0, 0, width, midpoint)).convert("RGB")
        bottom = source.crop((0, midpoint, width, height)).convert("RGB")
    top_path = output_dir / "chongwu-top.png"
    bottom_path = output_dir / "chongwu-bottom.png"
    top.save(top_path, format="PNG")
    bottom.save(bottom_path, format="PNG")
    return {
        "source": {
            "path": _relative(source_path, source_root or record_root),
            "sha256": sha256_file(source_path),
            "width": width,
            "height": height,
        },
        "rule": "split_at_floor_half_extra_row_to_bottom",
        "top": {
            "path": _relative(top_path, record_root),
            "sha256": sha256_file(top_path),
            "crop_box": [0, 0, width, midpoint],
            "width": width,
            "height": midpoint,
        },
        "bottom": {
            "path": _relative(bottom_path, record_root),
            "sha256": sha256_file(bottom_path),
            "crop_box": [0, midpoint, width, height],
            "width": width,
            "height": height - midpoint,
        },
        "uploaded_part": "bottom",
    }


def prepare_assets(config_path: Path, run_dir: Path) -> dict[str, Any]:
    config_path = config_path.resolve()
    run_dir = run_dir.resolve()
    config = json.loads(config_path.read_text(encoding="utf-8"))
    config_root = config_path.parent
    sources = config["sources"]
    style_workbook = (config_root / sources["style_workbook"]).resolve()
    composition_workbook = (config_root / sources["composition_workbook"]).resolve()
    character_sheet = (config_root / sources["character_sheet"]).resolve()
    for source in (style_workbook, composition_workbook, character_sheet):
        if not source.is_file():
            raise FileNotFoundError(source)

    cells = [item["style_cell"] for item in config["destinations"]]
    rows = sorted({item["composition_row"] for item in config["destinations"]})
    style_records = extract_style_images(
        style_workbook,
        sources["style_sheet"],
        cells,
        run_dir / "assets" / "styles",
        run_dir,
    )
    character = split_character_sheet(
        character_sheet,
        run_dir / "assets" / "character",
        run_dir,
        source_root=ROOT.parent,
    )
    return {
        "config": {
            "path": _relative(config_path, ROOT),
            "sha256": sha256_file(config_path),
        },
        "style_workbook": {
            "path": _relative(style_workbook, ROOT.parent),
            "sha256": sha256_file(style_workbook),
            "sheet": sources["style_sheet"],
        },
        "styles": style_records,
        "composition_workbook": {
            "path": _relative(composition_workbook, ROOT.parent),
            "sha256": sha256_file(composition_workbook),
            "sheet": sources["composition_sheet"],
        },
        "compositions": read_compositions(
            composition_workbook, sources["composition_sheet"], rows
        ),
        "character": character,
    }
