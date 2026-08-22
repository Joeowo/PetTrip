import importlib.util
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image


SCRIPT = Path(__file__).parents[1] / "scripts" / "issue12_assets.py"
spec = importlib.util.spec_from_file_location("issue12_assets", SCRIPT)
assets = importlib.util.module_from_spec(spec)
spec.loader.exec_module(assets)


def _png(path: Path, size=(20, 10), color="red") -> Path:
    Image.new("RGB", size, color).save(path)
    return path


def test_split_odd_height_gives_extra_row_to_bottom(tmp_path):
    source = _png(tmp_path / "source.png", size=(12, 7))

    record = assets.split_character_sheet(source, tmp_path / "out", tmp_path)

    assert record["top"]["height"] == 3
    assert record["bottom"]["height"] == 4
    assert record["top"]["crop_box"] == [0, 0, 12, 3]
    assert record["bottom"]["crop_box"] == [0, 3, 12, 7]
    assert len(record["source"]["sha256"]) == 64


def test_extract_style_image_by_exact_anchor(tmp_path):
    source = _png(tmp_path / "style.png")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "画风素材库"
    for column, value in enumerate(("画风", "游戏", "关键词", "详细提示词"), 1):
        sheet.cell(3, column).value = value
    image = WorksheetImage(source)
    sheet.add_image(image, "E3")
    workbook_path = tmp_path / "styles.xlsx"
    workbook.save(workbook_path)

    records = assets.extract_style_images(
        workbook_path, "画风素材库", ["E3"], tmp_path / "out", tmp_path
    )

    assert records[0]["cell"] == "E3"
    assert records[0]["width"] == 20
    assert records[0]["description"] == "详细提示词"
    assert records[0]["embedded_sha256"] == records[0]["output_sha256"]


def test_extract_style_image_fails_closed_for_missing_anchor(tmp_path):
    workbook = Workbook()
    workbook.active.title = "画风素材库"
    workbook_path = tmp_path / "styles.xlsx"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="E42 has 0"):
        assets.extract_style_images(
            workbook_path, "画风素材库", ["E42"], tmp_path / "out", tmp_path
        )


def test_read_composition_preserves_newlines(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "构图提示词"
    values = ["01", "分组", "名称", "电影", "特征", "第一行\n第二行", "提示"]
    for column, value in enumerate(values, 1):
        sheet.cell(9, column).value = value
    workbook_path = tmp_path / "composition.xlsx"
    workbook.save(workbook_path)

    record = assets.read_compositions(workbook_path, "构图提示词", [9])[0]

    assert record["fit"] == "第一行\n第二行"
